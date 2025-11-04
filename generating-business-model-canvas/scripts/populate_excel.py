#!/usr/bin/env python3
"""
Script principale per popolare il Business Model Canvas Excel.

Usage:
    python3 populate_excel.py <excel_file> <data_json_file>

JSON format:
{
  "metadata": {
    "project": "Nome Progetto",
    "author": "Team",
    "date": "2025-01-04",
    "version": "v1.0"
  },
  "sheet1": { ... },
  "sheet2": { ... },
  "sheet3": { ... },
  "sheet4": { ... }
}
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from sheet_populators import (
    populate_sheet1_business_model,
    populate_sheet2_lean_canvas,
    populate_sheet3_customer_personas,
    populate_sheet4_value_proposition
)


def populate_metadata(ws, metadata):
    """Popola metadata comuni a tutti i sheet"""
    ws['J4'] = metadata.get('project', '')
    ws['N4'] = metadata.get('author', '')
    ws['R4'] = metadata.get('date', '')
    ws['U4'] = metadata.get('version', 'v1.0')


def main():
    if len(sys.argv) != 3:
        print("Usage: python3 populate_excel.py <excel_file> <data_json_file>")
        print("\nExample:")
        print("  python3 populate_excel.py business-model-canvas.xlsx data.json")
        sys.exit(1)

    excel_file = sys.argv[1]
    data_file = sys.argv[2]

    # Verifica file esistano
    if not Path(excel_file).exists():
        print(f"❌ Error: Excel file not found: {excel_file}")
        sys.exit(1)

    if not Path(data_file).exists():
        print(f"❌ Error: Data file not found: {data_file}")
        sys.exit(1)

    # Carica dati JSON
    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"❌ Error: Invalid JSON in {data_file}: {e}")
        sys.exit(1)

    # Carica workbook
    try:
        wb = load_workbook(excel_file)
    except Exception as e:
        print(f"❌ Error: Cannot load Excel file: {e}")
        sys.exit(1)

    metadata = data.get('metadata', {})

    # Sheet 1: Business Model Canvas
    if 'sheet1' in data:
        try:
            ws = wb['Business Model Canvas']
            populate_metadata(ws, metadata)
            populate_sheet1_business_model(ws, data['sheet1'])
            print('✅ Sheet 1: Business Model Canvas completato')
        except Exception as e:
            print(f'⚠️  Sheet 1 error: {e}')

    # Sheet 2: Lean Canvas
    if 'sheet2' in data:
        try:
            ws = wb['Lean Canvas']
            populate_metadata(ws, metadata)
            populate_sheet2_lean_canvas(ws, data['sheet2'])
            print('✅ Sheet 2: Lean Canvas completato')
        except Exception as e:
            print(f'⚠️  Sheet 2 error: {e}')

    # Sheet 3: Customer Personas Canvas
    if 'sheet3' in data:
        try:
            ws = wb['Customer Personas Canvas']
            populate_metadata(ws, metadata)
            populate_sheet3_customer_personas(ws, data['sheet3'])
            print('✅ Sheet 3: Customer Personas Canvas completato')
        except Exception as e:
            print(f'⚠️  Sheet 3 error: {e}')

    # Sheet 4: Value Proposition Canvas i
    if 'sheet4' in data:
        try:
            ws = wb['Value Proposition Canvas i']
            populate_metadata(ws, metadata)
            populate_sheet4_value_proposition(ws, data['sheet4'])
            print('✅ Sheet 4: Value Proposition Canvas completato')
        except Exception as e:
            print(f'⚠️  Sheet 4 error: {e}')

    # Salva workbook
    try:
        wb.save(excel_file)
        print(f'\n✅ Excel salvato: {excel_file}')
        print(f'📊 4 canvas compilati')
    except Exception as e:
        print(f'❌ Error saving Excel: {e}')
        sys.exit(1)


if __name__ == '__main__':
    main()
