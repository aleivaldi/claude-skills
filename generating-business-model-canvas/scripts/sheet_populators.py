#!/usr/bin/env python3
"""
Funzioni per popolare ogni sheet del Business Model Canvas.

Ogni funzione riceve:
- ws: Worksheet openpyxl
- data: Dict con campi specifici per lo sheet

Tutti i campi sono OPZIONALI (usano .get() con default '').
Contenuti con \n per andare a capo in celle merged.
Priorità consigliate: 🔴 (critica), 🟡 (importante), 🟢 (nice-to-have)
"""

from openpyxl.styles import Alignment


def populate_sheet1_business_model(ws, data):
    """
    Popola Sheet 1: Business Model Canvas

    Args:
        ws: Worksheet openpyxl
        data: Dict con chiavi:
            - key_partners: str
            - key_activities: str
            - key_resources: str
            - value_propositions: str
            - customer_relationships: str
            - customer_segments: str
            - channels: str
            - cost_structure: str
            - revenue_streams: str
    """
    alignment = Alignment(wrap_text=True, vertical='top')

    # Celle merged - scrivi in top-left
    ws['B10'] = data.get('key_partners', '')
    ws['B10'].alignment = alignment

    ws['F10'] = data.get('key_activities', '')
    ws['F10'].alignment = alignment

    ws['F30'] = data.get('key_resources', '')
    ws['F30'].alignment = alignment

    ws['J10'] = data.get('value_propositions', '')
    ws['J10'].alignment = alignment

    ws['N10'] = data.get('customer_relationships', '')
    ws['N10'].alignment = alignment

    ws['R10'] = data.get('customer_segments', '')
    ws['R10'].alignment = alignment

    ws['N30'] = data.get('channels', '')
    ws['N30'].alignment = alignment

    ws['B50'] = data.get('cost_structure', '')
    ws['B50'].alignment = alignment

    ws['L50'] = data.get('revenue_streams', '')
    ws['L50'].alignment = alignment


def populate_sheet2_lean_canvas(ws, data):
    """
    Popola Sheet 2: Lean Canvas

    Args:
        ws: Worksheet openpyxl
        data: Dict con chiavi:
            - problem: str
            - existing_alternatives: str
            - solution: str
            - key_metrics: str
            - unique_value_proposition: str
            - high_level_concept: str
            - unfair_advantage: str
            - channels: str
            - customer_segments: str
            - early_adopters: str
            - cost_structure: str
            - revenue_streams: str
    """
    alignment = Alignment(wrap_text=True, vertical='top')

    ws['B10'] = data.get('problem', '')
    ws['B10'].alignment = alignment

    ws['B30'] = data.get('existing_alternatives', '')
    ws['B30'].alignment = alignment

    ws['F10'] = data.get('solution', '')
    ws['F10'].alignment = alignment

    ws['F30'] = data.get('key_metrics', '')
    ws['F30'].alignment = alignment

    ws['J10'] = data.get('unique_value_proposition', '')
    ws['J10'].alignment = alignment

    ws['J30'] = data.get('high_level_concept', '')
    ws['J30'].alignment = alignment

    ws['N10'] = data.get('unfair_advantage', '')
    ws['N10'].alignment = alignment

    ws['N30'] = data.get('channels', '')
    ws['N30'].alignment = alignment

    ws['R10'] = data.get('customer_segments', '')
    ws['R10'].alignment = alignment

    ws['R30'] = data.get('early_adopters', '')
    ws['R30'].alignment = alignment

    ws['B50'] = data.get('cost_structure', '')
    ws['B50'].alignment = alignment

    ws['L50'] = data.get('revenue_streams', '')
    ws['L50'].alignment = alignment


def populate_sheet3_customer_personas(ws, data):
    """
    Popola Sheet 3: Customer Personas Canvas

    Args:
        ws: Worksheet openpyxl
        data: Dict con chiavi:
            - personas: List[Dict] (max 3 personas)
                Ogni persona ha:
                - name: str
                - description: str
                - attributes: List[str] (attributi vari)
    """
    alignment = Alignment(wrap_text=True, vertical='top')

    personas = data.get('personas', [])

    # Colonne per persona: D, L, T
    persona_columns = ['D', 'L', 'T']

    for idx, persona in enumerate(personas[:3]):  # Max 3 personas
        col = persona_columns[idx]

        # Name
        ws[f'{col}8'] = persona.get('name', '')
        ws[f'{col}8'].alignment = alignment

        # Description
        ws[f'{col}10'] = persona.get('description', '')
        ws[f'{col}10'].alignment = alignment

        # Attributes (ogni 3 righe da riga 15)
        attributes = persona.get('attributes', [])
        row = 15
        for attr in attributes:
            if row <= 57:  # Limite righe
                ws[f'{col}{row}'] = attr
                ws[f'{col}{row}'].alignment = alignment
                row += 3


def populate_sheet4_value_proposition(ws, data):
    """
    Popola Sheet 4: Value Proposition Canvas

    Args:
        ws: Worksheet openpyxl
        data: Dict con chiavi:
            Product side:
            - benefits: str
            - features: str
            - value_proposition: str
            - experience: str
            Customer side:
            - wants: str
            - fears: str
            - rational_needs: str
            - substitutes: str
    """
    alignment = Alignment(wrap_text=True, vertical='top')

    # Product Side
    ws['B10'] = data.get('benefits', '')
    ws['B10'].alignment = alignment

    ws['B30'] = data.get('features', '')
    ws['B30'].alignment = alignment

    ws['F57'] = data.get('value_proposition', '')
    ws['F57'].alignment = alignment

    ws['F10'] = data.get('experience', '')
    ws['F10'].alignment = alignment

    # Customer Side
    ws['N10'] = data.get('wants', '')
    ws['N10'].alignment = alignment

    ws['R10'] = data.get('fears', '')
    ws['R10'].alignment = alignment

    ws['N30'] = data.get('rational_needs', '')
    ws['N30'].alignment = alignment

    ws['N54'] = data.get('substitutes', '')
    ws['N54'].alignment = alignment
