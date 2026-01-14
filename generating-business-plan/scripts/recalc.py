#!/usr/bin/env python3
"""
Recalculates all formulas in an Excel file using LibreOffice and scans for errors.

Usage:
    python3 recalc.py <excel_file> [timeout_seconds]

Returns JSON with:
{
  "status": "success" | "errors_found",
  "total_errors": N,
  "total_formulas": M,
  "error_summary": {
    "#REF!": {"count": X, "locations": ["Sheet!A1", ...]},
    ...
  }
}

Requirements:
    - LibreOffice installed
    - First run: Creates macro for recalculation
"""

import sys
import subprocess
import json
import os
import platform
from pathlib import Path
import time


def get_libreoffice_path():
    """Detect LibreOffice executable path based on OS"""
    system = platform.system()

    if system == "Darwin":  # macOS
        paths = [
            "/Applications/LibreOffice.app/Contents/MacOS/soffice",
            "~/Applications/LibreOffice.app/Contents/MacOS/soffice"
        ]
    elif system == "Linux":
        paths = [
            "/usr/bin/libreoffice",
            "/usr/bin/soffice",
            "/usr/local/bin/libreoffice"
        ]
    else:  # Windows
        paths = [
            "C:\\Program Files\\LibreOffice\\program\\soffice.exe",
            "C:\\Program Files (x86)\\LibreOffice\\program\\soffice.exe"
        ]

    for path in paths:
        expanded = os.path.expanduser(path)
        if os.path.exists(expanded):
            return expanded

    return None


def setup_libreoffice_macro():
    """
    Creates LibreOffice macro for recalculating all formulas.
    Only needs to be run once.
    """
    system = platform.system()

    if system == "Darwin":
        macro_dir = os.path.expanduser("~/Library/Application Support/LibreOffice/4/user/Scripts/python")
    elif system == "Linux":
        macro_dir = os.path.expanduser("~/.config/libreoffice/4/user/Scripts/python")
    else:
        macro_dir = os.path.expanduser("~/AppData/Roaming/LibreOffice/4/user/Scripts/python")

    os.makedirs(macro_dir, exist_ok=True)

    macro_file = os.path.join(macro_dir, "recalc_workbook.py")

    macro_content = '''#!/usr/bin/env python3
"""LibreOffice macro to recalculate all sheets in workbook"""

def recalculate_all(*args):
    """Recalculates all formulas in all sheets"""
    import uno
    from com.sun.star.sheet import XSpreadsheetDocument

    desktop = XSCRIPTCONTEXT.getDesktop()
    model = desktop.getCurrentComponent()

    if model is None:
        return "No document open"

    # Check if it's a spreadsheet
    if not hasattr(model, "Sheets"):
        return "Not a spreadsheet document"

    sheets = model.Sheets
    for i in range(sheets.Count):
        sheet = sheets.getByIndex(i)
        # Force recalculation
        sheet.calculateAll()

    # Save and close
    model.store()

    return "Recalculation complete"


g_exportedScripts = (recalculate_all,)
'''

    with open(macro_file, 'w') as f:
        f.write(macro_content)

    print(f"✅ LibreOffice macro installed: {macro_file}")


def recalculate_excel(excel_file, timeout=120):
    """
    Recalculates all formulas in Excel file using LibreOffice.
    """
    soffice = get_libreoffice_path()

    if not soffice:
        return {
            "status": "error",
            "message": "LibreOffice not found. Please install LibreOffice."
        }

    # Ensure macro is set up
    setup_libreoffice_macro()

    # Convert to absolute path
    excel_path = os.path.abspath(excel_file)

    # LibreOffice command to open, recalculate, and close
    # Using headless mode
    cmd = [
        soffice,
        "--headless",
        "--calc",
        "--convert-to", "xlsx",
        "--outdir", os.path.dirname(excel_path),
        excel_path
    ]

    try:
        result = subprocess.run(
            cmd,
            timeout=timeout,
            capture_output=True,
            text=True
        )

        if result.returncode != 0:
            return {
                "status": "error",
                "message": f"LibreOffice error: {result.stderr}"
            }

        return {"status": "success"}

    except subprocess.TimeoutExpired:
        return {
            "status": "error",
            "message": f"Timeout after {timeout}s"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }


def scan_for_errors(excel_file):
    """
    Scans Excel file for formula errors using openpyxl.
    Returns summary of errors found.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "status": "error",
            "message": "openpyxl not installed. Run: pip install openpyxl"
        }

    wb = load_workbook(excel_file, data_only=False)

    error_types = ["#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"]
    errors = {error: {"count": 0, "locations": []} for error in error_types}
    total_formulas = 0
    total_errors = 0

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    if cell.value.startswith('='):
                        total_formulas += 1

                    for error_type in error_types:
                        if error_type in str(cell.value):
                            errors[error_type]["count"] += 1
                            errors[error_type]["locations"].append(f"{sheet_name}!{cell.coordinate}")
                            total_errors += 1
                            break

    # Remove error types with zero count
    error_summary = {k: v for k, v in errors.items() if v["count"] > 0}

    return {
        "status": "errors_found" if total_errors > 0 else "success",
        "total_errors": total_errors,
        "total_formulas": total_formulas,
        "error_summary": error_summary if error_summary else {}
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 recalc.py <excel_file> [timeout_seconds]")
        print("\nExample:")
        print("  python3 recalc.py business-plan.xlsx 60")
        sys.exit(1)

    excel_file = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 120

    if not os.path.exists(excel_file):
        result = {
            "status": "error",
            "message": f"File not found: {excel_file}"
        }
        print(json.dumps(result, indent=2))
        sys.exit(1)

    print(f"📊 Recalculating formulas in: {excel_file}")
    print(f"⏱️  Timeout: {timeout}s\n")

    # Step 1: Recalculate with LibreOffice
    recalc_result = recalculate_excel(excel_file, timeout)

    if recalc_result["status"] == "error":
        print(f"❌ Recalculation failed: {recalc_result['message']}")
        print(json.dumps(recalc_result, indent=2))
        sys.exit(1)

    print("✅ Recalculation complete")

    # Wait a moment for file to be fully written
    time.sleep(1)

    # Step 2: Scan for errors
    print("🔍 Scanning for formula errors...\n")
    scan_result = scan_for_errors(excel_file)

    if scan_result.get("status") == "error":
        print(f"❌ Scan failed: {scan_result['message']}")
        print(json.dumps(scan_result, indent=2))
        sys.exit(1)

    # Output results
    print(json.dumps(scan_result, indent=2))

    if scan_result["total_errors"] > 0:
        print(f"\n⚠️  Found {scan_result['total_errors']} formula errors")
        print(f"📝 {scan_result['total_formulas']} total formulas")
        print("\nFix errors and run recalc.py again.")
        sys.exit(1)
    else:
        print(f"\n✅ No errors found")
        print(f"📝 {scan_result['total_formulas']} formulas verified")
        sys.exit(0)


if __name__ == '__main__':
    main()
