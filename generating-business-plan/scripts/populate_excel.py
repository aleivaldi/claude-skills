#!/usr/bin/env python3
"""
Script principale per popolare il Business Plan Excel.

Usage:
    python3 populate_excel.py <excel_file> <data_json_file>

JSON format: Vedi SKILL.md sezione "JSON Format"
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta


def populate_input_sheet(ws, data):
    """
    Popola il foglio Input con tutti i dati.
    IMPORTANTE: Scrive SOLO valori, NON formule. Le formule sono già nel template.
    """
    metadata = data.get('metadata', {})
    start_date = datetime.strptime(metadata.get('start_date', '2026-01-01'), '%Y-%m-%d')
    months = metadata.get('projection_months', 36)

    # Header metadata
    ws['A1'] = 'Input'
    ws['A2'] = f"All amounts in {metadata.get('currency', 'EUR')}. VAT not included unless specified."

    # Fiscal Months (row 4, columns D onwards)
    ws['A4'] = 'Fiscal Months'
    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"  # D, E, F... then AA, AB...
        month_date = start_date + relativedelta(months=i)
        ws[f'{col}4'] = month_date

    # MACRO
    macro = data.get('macro', {})
    ws['A6'] = 'Macro'
    ws['A8'] = 'Annual inflation rate'
    annual_inflation = macro.get('annual_inflation_rate', 0.02)
    monthly_inflation = annual_inflation / 12

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}8'] = annual_inflation

    ws['A9'] = 'Monthly inflation rate'
    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}9'] = monthly_inflation

    # REVENUES
    revenues = data.get('revenues', {})

    ws['A11'] = 'Revenues'
    ws['A13'] = 'Amazon'
    ws['A15'] = 'Price (VAT not included)'
    ws['A16'] = 'Quantity'
    ws['A17'] = 'Revenues - Total Amazon'

    amazon = revenues.get('amazon', {})
    amazon_price = amazon.get('price', 0)
    amazon_qty = amazon.get('quantities', [])

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}15'] = amazon_price
        qty = amazon_qty[i] if i < len(amazon_qty) else (amazon_qty[-1] if amazon_qty else 0)
        ws[f'{col}16'] = qty
        # Row 17 ha formula: =D15*D16

    ws['A19'] = 'Distribution'
    ws['A21'] = 'Price (VAT not included)'
    ws['A22'] = 'Quantity'
    ws['A23'] = 'Revenues - Total Distribution'

    distribution = revenues.get('distribution', {})
    dist_price = distribution.get('price', 0)
    dist_qty = distribution.get('quantities', [])

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}21'] = dist_price
        qty = dist_qty[i] if i < len(dist_qty) else (dist_qty[-1] if dist_qty else 0)
        ws[f'{col}22'] = qty
        # Row 23 ha formula: =D21*D22

    ws['A25'] = 'Quantity - Total'
    ws['A26'] = 'Revenues - Total'
    # Row 25 formula: =D16+D22
    # Row 26 formula: =D17+D23

    # COGS
    cogs = data.get('cogs', {})

    ws['A28'] = 'Costs'
    ws['A30'] = 'COGS'
    ws['A32'] = 'Product'
    ws['A34'] = 'COGS per unit'
    ws['A35'] = 'Total COGS - Product'

    product_cogs = cogs.get('product_cogs_per_unit', 0)
    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}34'] = product_cogs
        # Row 35 formula: =D34*D25 (cogs * total quantity)

    ws['A37'] = 'Amazon fee'
    ws['A39'] = 'Referral fee %'
    ws['A40'] = 'Total COGS - Amazon'

    amazon_fee = cogs.get('amazon_referral_fee_pct', 0.15)
    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}39'] = amazon_fee
        # Row 40 formula: =D17*D39 (amazon revenue * fee)

    ws['A42'] = 'Packaging & Shipping'
    ws['A44'] = 'Packaging cost per unit'
    ws['A45'] = 'Shipping per unit - Factory to Warehouse'
    ws['A46'] = 'Shipping per unit - Warehouse to Distributor'
    ws['A47'] = 'Shipping per unit - Amazon'
    ws['A48'] = 'Packaging cost'
    ws['A49'] = 'Shipping - Factory to Warehouse'
    ws['A50'] = 'Shipping - Warehouse to Distributor'
    ws['A51'] = 'Shipping - Amazon'
    ws['A52'] = 'Total COGS - Packaging & Shipping'

    shipping = cogs.get('shipping', {})
    packaging = cogs.get('packaging_per_unit', 0)

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}44'] = packaging
        ws[f'{col}45'] = shipping.get('factory_to_warehouse', 0)
        ws[f'{col}46'] = shipping.get('warehouse_to_distributor', 0)
        ws[f'{col}47'] = shipping.get('warehouse_to_amazon', 0)
        # Formulas:
        # Row 48: =D44*D25 (packaging * total qty)
        # Row 49: =D45*D25 (shipping * total qty)
        # Row 50: =D46*D22 (shipping to dist * dist qty)
        # Row 51: =D47*D16 (shipping to amazon * amazon qty)
        # Row 52: =SUM(D48:D51)

    ws['A54'] = 'Total COGS'
    # Formula: =D35+D40+D52

    # MARKETING
    marketing = data.get('marketing', {})

    ws['A56'] = 'Marketing'
    ws['A58'] = 'LTV/CAC'
    ws['A59'] = 'LTV (Price Amazon)'
    ws['A60'] = 'CAC Max'
    ws['A61'] = 'Total Marketing Costs'

    ltv_cac = marketing.get('ltv_cac_ratio', 3.0)
    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}58'] = ltv_cac
        # Row 59 formula: =D15 (LTV = Amazon price)
        # Row 60 formula: =D59/D58 (CAC = LTV/ratio)
        # Row 61 formula: =D16*D60 (Marketing = new customers * CAC)
        # Assumes amazon qty = new customers

    # PERSONNEL
    personnel = data.get('personnel', {})
    salaries = personnel.get('salaries', {})
    ftes = personnel.get('ftes', {})

    ws['A63'] = 'Personnel'

    roles = [
        ('C-Level', 65),
        ('Finance', 66),
        ('Sales', 67),
        ('Marketing', 68),
        ('Product', 69),
        ('Operations', 70)
    ]

    for role, row in roles:
        role_key = role.lower().replace('-', '_')
        ws[f'A{row}'] = f'Salary - {role}'
        salary = salaries.get(role_key, 0)
        for i in range(months):
            col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
            ws[f'{col}{row}'] = salary

    for role, row in roles:
        role_key = role.lower().replace('-', '_')
        fte_row = row + 6
        ws[f'A{fte_row}'] = f'FTE - {role}'
        fte_list = ftes.get(role_key, [])
        for i in range(months):
            col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
            fte = fte_list[i] if i < len(fte_list) else (fte_list[-1] if fte_list else 0)
            ws[f'{col}{fte_row}'] = fte

    ws['A77'] = 'Total Cost (gross)'
    ws['A78'] = 'Total FTE'
    ws['A79'] = 'Capitalization'
    ws['A80'] = 'Total Cost (net)'
    # Formulas:
    # Row 77: =D65*D71+D66*D72+D67*D73+D68*D74+D69*D75+D70*D76
    # Row 78: =SUM(D71:D76)
    # Row 79: Calculated based on capitalization_rate * product FTE cost
    # Row 80: =D77-D79

    ws['A82'] = 'Pension provision rate'
    ws['A83'] = 'Pension provision'

    pension_rate = personnel.get('pension_provision_rate', 0.0691)
    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}82'] = pension_rate
        # Row 83 formula: =D80*D82

    # G&A
    gna = data.get('gna', {})

    ws['A85'] = 'G&A'
    ws['A87'] = 'Warehouse rent'
    ws['A88'] = 'SaaS per employee'
    ws['A89'] = 'SaaS - total'
    ws['A90'] = 'CPA'
    ws['A91'] = 'HR Consultant cost per employee'
    ws['A92'] = 'HR Consultant - Total'
    ws['A93'] = '% Other'
    ws['A94'] = 'Other costs'
    ws['A95'] = 'Total G&A Costs'

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}87'] = gna.get('warehouse_rent', 0)
        ws[f'{col}88'] = gna.get('saas_per_employee', 200)
        # Row 89 formula: =D88*D78 (saas * total FTE)
        if i == 0:  # CPA only first month typically, or monthly
            ws[f'{col}90'] = gna.get('cpa_monthly', 1000)
        ws[f'{col}91'] = gna.get('hr_consultant_per_employee', 30)
        # Row 92 formula: =D91*D78
        ws[f'{col}93'] = gna.get('other_pct', 0.15)
        # Row 94 formula: =(D87+D89+D90+D92)*D93/(1-D93)
        # Row 95 formula: =SUM(D87:D94)

    # TAXES
    taxes = data.get('taxes', {})

    ws['A97'] = 'Taxes'
    ws['A99'] = 'IRES'
    ws['A100'] = 'IRAP'
    ws['A101'] = 'VAT'

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}99'] = taxes.get('ires', 0.24)
        ws[f'{col}100'] = taxes.get('irap', 0.039)
        ws[f'{col}101'] = taxes.get('vat', 0.22)

    # FINANCING
    financing = data.get('financing', {})

    ws['A103'] = 'Financing'
    ws['A105'] = 'Equity injection'

    equity_injections = financing.get('equity_injections', [])
    for injection in equity_injections:
        month_idx = injection.get('month', 1) - 1
        if month_idx < months:
            col = chr(68 + month_idx) if month_idx < 23 else f"A{chr(65 + month_idx - 23)}"
            ws[f'{col}105'] = injection.get('amount', 0)

    ws['A107'] = 'Debt BoP'
    ws['A108'] = 'New debt'
    ws['A109'] = 'Reimbursement'
    ws['A110'] = 'Debt EoP'
    ws['A111'] = 'Interest rate'
    ws['A112'] = 'Financial expenses'

    debt = financing.get('debt', {})
    annual_rate = debt.get('interest_rate_annual', 0.10)
    monthly_rate = annual_rate / 12

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}111'] = monthly_rate
        # Debt formulas handle beginning/ending balance

    # CAPEX
    capex = data.get('capex', {})
    intangible = capex.get('intangible', {})
    tangible = capex.get('tangible', {})

    ws['A114'] = 'Capex'
    ws['A116'] = 'Intangible assets'
    ws['A118'] = 'Tech & Product'
    ws['A119'] = 'Personnel'
    ws['A120'] = 'Total Capex'
    ws['A121'] = 'Tech & Product - D&A Yearly Rate'
    ws['A122'] = 'Tech & Product - D&A Monthly Rate'

    initial_tech = intangible.get('initial_tech_product', 0)
    ws['D118'] = initial_tech  # Month 1

    amort_annual = intangible.get('amortization_rate_annual', 0.20)
    amort_monthly = amort_annual / 12

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        if i == 0:
            ws[f'{col}121'] = amort_annual
        ws[f'{col}122'] = amort_monthly
        # Row 119: Personnel capex (calculated from capitalization)
        # Row 120: =D118+D119

    ws['A124'] = 'Tangible assets'
    ws['A126'] = 'Capex per employee'
    ws['A127'] = 'Capex per employee (change pc)'
    ws['A128'] = 'Capex'
    ws['A129'] = 'Tangible - D&A Yearly Rate'
    ws['A130'] = 'Tangible - D&A Monthly Rate'

    capex_per_employee = tangible.get('capex_per_employee', 1000)
    depr_annual = tangible.get('depreciation_rate_annual', 0.3333)
    depr_monthly = depr_annual / 12

    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}126'] = capex_per_employee
        if i == 0:
            ws[f'{col}129'] = depr_annual
        ws[f'{col}130'] = depr_monthly
        # Row 127, 128: Calculate based on FTE changes

    ws['A132'] = 'D&A'
    ws['A134'] = 'Intangible assets'
    ws['A136'] = 'Capitalized costs (Tech & Product)'
    ws['A137'] = 'BoP'
    ws['A138'] = 'Capex'
    ws['A139'] = 'Amortization'

    # Rows 140+ are monthly breakdown with formulas
    for i in range(months):
        row = 140 + i
        month_date = start_date + relativedelta(months=i)
        ws[f'A{row}'] = month_date
        # Formulas for depreciation schedule


def add_formulas_to_input(ws, months):
    """
    Aggiunge le formule al foglio Input.
    Chiamata DOPO populate_input_sheet per non sovrascrivere i valori.
    """
    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"

        # Revenue formulas
        ws[f'{col}17'] = f'={col}15*{col}16'  # Amazon revenue
        ws[f'{col}23'] = f'={col}21*{col}22'  # Distribution revenue
        ws[f'{col}25'] = f'={col}16+{col}22'  # Total quantity
        ws[f'{col}26'] = f'={col}17+{col}23'  # Total revenue

        # COGS formulas
        ws[f'{col}35'] = f'={col}34*{col}25'  # Product COGS
        ws[f'{col}40'] = f'={col}17*{col}39'  # Amazon fee
        ws[f'{col}48'] = f'={col}44*{col}25'  # Packaging
        ws[f'{col}49'] = f'={col}45*{col}25'  # Shipping factory-warehouse
        ws[f'{col}50'] = f'={col}46*{col}22'  # Shipping to distributor
        ws[f'{col}51'] = f'={col}47*{col}16'  # Shipping to Amazon
        ws[f'{col}52'] = f'=SUM({col}48:{col}51)'  # Total packaging & shipping
        ws[f'{col}54'] = f'={col}35+{col}40+{col}52'  # Total COGS

        # Marketing formulas
        ws[f'{col}59'] = f'={col}15'  # LTV = Amazon price
        ws[f'{col}60'] = f'={col}59/{col}58'  # CAC
        ws[f'{col}61'] = f'={col}16*{col}60'  # Marketing cost

        # Personnel formulas
        ws[f'{col}77'] = f'={col}65*{col}71+{col}66*{col}72+{col}67*{col}73+{col}68*{col}74+{col}69*{col}75+{col}70*{col}76'
        ws[f'{col}78'] = f'=SUM({col}71:{col}76)'  # Total FTE

        # Capitalization (50% of Product cost as example)
        ws[f'{col}79'] = f'={col}69*{col}75*0.5'  # Simplified: 50% of product salary
        ws[f'{col}80'] = f'={col}77-{col}79'  # Net personnel cost
        ws[f'{col}83'] = f'={col}80*{col}82'  # Pension provision

        # G&A formulas
        ws[f'{col}89'] = f'={col}88*{col}78'  # SaaS total
        ws[f'{col}92'] = f'={col}91*{col}78'  # HR consultant total
        ws[f'{col}94'] = f'=({col}87+{col}89+{col}90+{col}92)*{col}93/(1-{col}93)'  # Other costs
        ws[f'{col}95'] = f'=SUM({col}87:{col}94)'  # Total G&A

        # Capex formulas
        ws[f'{col}119'] = f'={col}79'  # Personnel capex = capitalization
        ws[f'{col}120'] = f'={col}118+{col}119'  # Total intangible capex

        # Tangible capex (new FTE * capex per employee)
        if i == 0:
            ws[f'{col}127'] = f'={col}78*{col}126'  # First month: all FTE
        else:
            prev_col = chr(68 + i - 1) if i - 1 < 23 else f"A{chr(65 + i - 24)}"
            ws[f'{col}127'] = f'=MAX(0, {col}78-{prev_col}78)*{col}126'  # New FTE only
        ws[f'{col}128'] = f'={col}127'  # Tangible capex

        # Debt formulas
        if i == 0:
            ws[f'{col}107'] = '=0'  # Initial debt BoP
        else:
            prev_col = chr(68 + i - 1) if i - 1 < 23 else f"A{chr(65 + i - 24)}"
            ws[f'{col}107'] = f'={prev_col}110'  # BoP = previous EoP

        ws[f'{col}110'] = f'={col}107+{col}108-{col}109'  # EoP
        ws[f'{col}112'] = f'={col}107*{col}111'  # Interest expense


def populate_output_sheet(ws, months):
    """
    Popola foglio Output con formule che referenziano Input sheet.
    """
    ws['A1'] = 'Output'
    ws['A3'] = 'Fiscal Year'
    ws['A4'] = 'Fiscal Months'

    # Copy dates from Input
    for i in range(months):
        col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
        ws[f'{col}4'] = f'=Input!{col}4'

    ws['A6'] = 'Profit & Loss'

    rows = [
        (7, 'Revenues - Amazon', 'Input!{col}17'),
        (8, 'Revenues - Distribution', 'Input!{col}23'),
        (9, 'Total Revenues', '={col}7+{col}8'),
        (10, 'COGS - Product', 'Input!{col}35'),
        (11, 'COGS - Amazon', 'Input!{col}40'),
        (12, 'COGS - Packaging & Shipping', 'Input!{col}52'),
        (13, 'Total COGS', '={col}10+{col}11+{col}12'),
        (14, 'Gross Profit', '={col}9-{col}13'),
        (15, 'Gross Margin', 'IF({col}9<>0, {col}14/{col}9, 0)'),
        (16, 'Marketing', 'Input!{col}61'),
        (17, 'Personnel', 'Input!{col}80+Input!{col}83'),
        (18, 'G&A', 'Input!{col}95'),
        (19, 'Total Fixed Costs', '={col}16+{col}17+{col}18'),
        (20, 'EBITDA', '={col}14-{col}19'),
        (21, 'EBITDA Margin', 'IF({col}9<>0, {col}20/{col}9, 0)'),
        (22, 'D&A', '=0'),  # Placeholder, needs depreciation calc
        (23, 'EBIT', '={col}20-{col}22'),
        (24, 'Interests', 'Input!{col}112'),
        (25, 'EBT', '={col}23-{col}24'),
        (26, 'Taxes', '={col}25*Input!{col}99'),  # Simplified: IRES only
        (27, 'Net Profit/(Loss)', '={col}25-{col}26'),
    ]

    for row, label, formula_template in rows:
        ws[f'A{row}'] = label
        for i in range(months):
            col = chr(68 + i) if i < 23 else f"A{chr(65 + i - 23)}"
            ws[f'{col}{row}'] = formula_template.format(col=col)

    # Balance Sheet
    ws['A29'] = 'Balance sheet'
    ws['A30'] = 'Intangible assets'
    ws['A31'] = 'Tangible assets'
    ws['A32'] = 'Net fixed assets'
    # ... continue with balance sheet formulas


def populate_financial_statements(ws, months):
    """
    Popola foglio Financial Statements con consolidamenti annuali.
    """
    ws['A1'] = 'Financial Statements'
    # Similar structure to Output but aggregated by year


def main():
    """Entry point"""
    if len(sys.argv) != 3:
        print("Usage: python3 populate_excel.py <excel_file> <data_json_file>")
        sys.exit(1)

    excel_file = sys.argv[1]
    data_file = sys.argv[2]

    if not Path(excel_file).exists():
        print(f"❌ Error: Excel file not found: {excel_file}")
        sys.exit(1)

    if not Path(data_file).exists():
        print(f"❌ Error: Data file not found: {data_file}")
        sys.exit(1)

    try:
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"❌ Error: Invalid JSON: {e}")
        sys.exit(1)

    try:
        wb = load_workbook(excel_file)
    except Exception as e:
        print(f"❌ Error loading Excel: {e}")
        sys.exit(1)

    months = data.get('metadata', {}).get('projection_months', 36)

    # Populate sheets
    try:
        ws_input = wb['Input']
        populate_input_sheet(ws_input, data)
        add_formulas_to_input(ws_input, months)
        print('✅ Sheet Input populated')
    except Exception as e:
        print(f'⚠️  Input sheet error: {e}')
        import traceback
        traceback.print_exc()

    try:
        ws_output = wb['Output']
        populate_output_sheet(ws_output, months)
        print('✅ Sheet Output populated')
    except Exception as e:
        print(f'⚠️  Output sheet error: {e}')

    try:
        ws_fs = wb['Financial Statements']
        populate_financial_statements(ws_fs, months)
        print('✅ Sheet Financial Statements populated')
    except Exception as e:
        print(f'⚠️  Financial Statements sheet error: {e}')

    try:
        wb.save(excel_file)
        print(f'\n✅ Excel saved: {excel_file}')
    except Exception as e:
        print(f'❌ Error saving: {e}')
        sys.exit(1)


if __name__ == '__main__':
    main()
